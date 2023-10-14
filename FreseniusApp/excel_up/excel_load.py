
from sys import stdout
import openpyxl
from ..models import WorkUser,WorkDate
import datetime
from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from datetime import datetime
from django.utils import timezone

def excelLoad(files):
    model_write(files)

 

      

def model_write(files):
    print(files)
    excel_file = openpyxl.load_workbook(files)
    ws = excel_file.active
    max_column=ws.max_column
    max_row = ws.max_row
    
    work=list()
    for value in ws.iter_cols(min_row=1, max_col=max_column, max_row=max_row,min_col=1):
        work_list=list()
        for cell in value:
            # print(cell.value)
            values=cell.value
            if values !=None:
                print(values)
                work_list.append(values)
        work.append(work_list)
    

  
